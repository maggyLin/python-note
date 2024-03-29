from docx import Document
import openpyxl

# 獲取 test plan data =============================================

document = Document("Test Plan_demo.docx")
title_key = "Test Items Command"
key_table_data = []


# 抓取所有tables
tables = document.tables
for i in range(0,len(tables)):
    table = tables[i]
    # 抓取有title_key tables
    title_str = table.cell(0,0).text
    # print(title_str)
    if(title_str.find(title_key) != -1):
        table_list = [['' for i in range(len(table.columns))] for j in range(len(table.rows))]
        for j in range(0,len(table.rows)):
            for k in range(0,len(table.columns)):
                table_list[j][k] = table.cell(j,k).text  
        # print(table_list)
        key_table_data.append(table_list)
        

# 建立 Excel =================================================
wb = openpyxl.Workbook()


for i in range(0,len(key_table_data)):
    table_data = key_table_data[i]
    sheet = wb.create_sheet("test", 0)
    my_sheet = wb.active
    for j in range(0,len(table_data)):
        for k in range(0,len(table_data[j])):
            # print(table_data[j][k])
            my_sheet.cell(row=(j+1), column=(k+1)).value = table_data[j][k]

wb.save("sample_data3.xlsx")